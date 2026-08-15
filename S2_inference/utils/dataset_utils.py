import pyvista
import numpy as np
import random
import matplotlib.pyplot as plt
import neurokit2 as nk
import os
import pandas as pd
import scipy.signal
import glob
from matplotlib.widgets import Button
from PIL import Image
from collections import defaultdict
import re
import openpyxl
from datetime import datetime, timedelta
import shutil
import pymupdf
import csv
# ------------------------------MRI------------------------------- #
def getCobiveco_vtu(cobiveco_fileName): # Read Cobiveco data in .vtu format
    cobiveco_vol = pyvista.read(cobiveco_fileName) #, force_ext='.vtu'

    cobiveco_nodesXYZ = cobiveco_vol.points
    cobiveco_nodes_array = cobiveco_vol.point_data
    # Apex-to-Base - ab
    ab = cobiveco_nodes_array['ab']
    # Rotation angle - rt
    rt = cobiveco_nodes_array['rt']
    # Transmurality - tm
    tm = cobiveco_nodes_array['tm']
    # Ventricle - tv
    tv = cobiveco_nodes_array['tv']
    ### Add more coordinates (Mengxiao, 20260109)
    # posterior-to-anterior (Projection) - posterior=0, anterior=1
    # aprt = cobiveco_nodes_array['aprt']
    # # lv-to-rv (Projection) - LV=0, RV=1
    # rvlv = cobiveco_nodes_array['rvlv']
    # AHA-17 map - aha
    aha = cobiveco_nodes_array['aha']

    return cobiveco_nodesXYZ, np.transpose(np.array([ab, rt, tm, tv, aha], dtype=float))

def getScarLabel_vtu(scar_fileName): # Read scar label in .vtu format
    scar_vol = pyvista.read(scar_fileName) #, force_ext='.vtu'
    scar_nodes_array = scar_vol.point_data
    scar_label = scar_nodes_array['ScarLabel']
    return np.array(scar_label, dtype=int)

###　point cloud augmentation　###
# translate point cloud
def translate_point(point):
    point = np.array(point)
    shift = [random.uniform(-0.5, 0.5), random.uniform(-0.5, 0.5), random.uniform(-0.5, 0.5)]
    shift = np.expand_dims(np.array(shift), axis=0)
    shifted_point = np.repeat(shift, point.shape[0], axis=0)
    shifted_point += point

    return shifted_point

# add Gaussian noise 
def jitter_point(point, sigma=0.01, clip=0.01):
    assert(clip > 0)
    point = np.array(point)
    point = point.reshape(-1,3)
    Row, Col = point.shape
    jittered_point = np.clip(sigma * np.random.randn(Row, Col), -1*clip, clip)
    jittered_point += point

    return jittered_point

# rotate point cloud
def rotate_point(point, rotation_angle=0.5*np.pi):
    point = np.array(point)
    cos_theta = np.cos(rotation_angle)
    sin_theta = np.sin(rotation_angle)
    # Rotation around X axis
    rotation_matrix_X = np.array([[1, 0, 0],
                                [0, cos_theta, -sin_theta],
                                [0, sin_theta, cos_theta]])
    # Rotation around Y axis
    rotation_matrix_Y = np.array([[cos_theta, 0, sin_theta],
                                [0, 1, 0],
                                [-sin_theta, 0, cos_theta]])
    # Rotation around Z axis
    rotation_matrix_Z = np.array([[cos_theta, sin_theta, 0],
                                [-sin_theta, cos_theta, 0],
                                [0, 0, 1]])   

    rotated_point = np.dot(point.reshape(-1, 3), rotation_matrix_Z)

    return rotated_point

# normalize point cloud based on apex coordinate
def normalize_data(PC, Coord_apex):
    """ Normalize the point cloud, use coordinates of centroid/ apex,
        Input:
            NxC array
        Output:
            NxC array
    """
    N, C = PC.shape
    normal_data = np.zeros((N, C))
    # centroid = np.mean(PC, axis=0)
    PC = PC - Coord_apex
    # m = np.max(np.sqrt(np.sum(PC ** 2, axis=1)))
    # PC = PC / m
    # normal_data = PC

    # compute the minimum and maximum values of each coordinate
    min_coords = np.min(PC, axis=0)
    max_coords = np.max(PC, axis=0)

    # normalize the point cloud coordinates
    normal_data = (PC - min_coords) / (max_coords - min_coords)

    return normal_data

def resample_pcd_ATM(pcd, ATM, n):
    """Drop or duplicate points so that pcd has exactly n points"""
    idx_root_nodes = np.where(ATM[:, 0]==1.0) # ATM[:, 0]
    prob = 1/(pcd.shape[0]-idx_root_nodes[0].shape[0])
    node_prob = prob*np.ones(pcd.shape[0])
    node_prob[idx_root_nodes] = 0
    idx = np.random.choice(np.arange(pcd.shape[0]), n-idx_root_nodes[0].shape[0], p=node_prob, replace=False)
    idx_remained = np.union1d(idx, idx_root_nodes)
    # idx_updated_permuted = np.random.permutation(idx_updated)
    # if idx_updated_permuted.shape[0] < n:
    #     idx = np.concatenate([idx, np.random.randint(pcd.shape[0], size=n-pcd.shape[0])])
    
    return pcd[idx_remained], ATM[idx_remained], idx_remained

def resample_pcd_ATM_ori(pcd, ATM, n):
    """Drop or duplicate points so that pcd has exactly n points"""
    idx = np.random.permutation(pcd.shape[0])
    if idx.shape[0] < n:
        idx = np.concatenate([idx, np.random.randint(pcd.shape[0], size=n-pcd.shape[0])])
    return pcd[idx[:n]], ATM[idx[:n]]

def resample_gd(gt_output, num_coarse, num_dense): #added by Lei in 2022/02/10 to seperately resample groundtruth label
    """Drop or duplicate points so that pcd has exactly n points"""
    choice = np.random.choice(len(gt_output), num_coarse, replace=True)
    coarse_gt = gt_output[choice, :]
    dense_gt = resample_pcd(gt_output, num_dense)
    return coarse_gt, dense_gt

def resample_pcd(pcd, n):
    """Drop or duplicate points so that pcd has exactly n points"""
    idx = np.random.permutation(pcd.shape[0])
    if idx.shape[0] < n:
        idx = np.concatenate([idx, np.random.randint(pcd.shape[0], size=n-pcd.shape[0])])
    return pcd[idx[:n]], idx[:n]


# ------------------------------MRI end------------------------------- #

# ------------------------------ECG start------------------------------- #


class InteractiveSelector:
    """
    Let users click a Matplotlib plot to select or correct multiple feature-point positions.
    Left click: move the nearest point.
    Right click: add a point.
    """

    def __init__(self, ax, signal, initial_peak_idx=None,
                 title="Click to select/correct QRS Onsets"):
        self.ax = ax
        self.signal = signal

        # ---- Core change 1: normalize to a list. ----
        if initial_peak_idx is None:
            self.selected_indices = []
        elif isinstance(initial_peak_idx, (list, tuple, np.ndarray)):
            self.selected_indices = list(initial_peak_idx)
        else:
            self.selected_indices = [int(initial_peak_idx)]

        self.ax.set_title(f"{title}. Close window to confirm.")
        self.line, = self.ax.plot(signal, label='Filtered Signal')

        # ---- Core change 2: let the marker support multiple points. ----
        self.marker, = self.ax.plot(
            [], [],
            marker='o', markersize=8, color='red',
            linestyle='', label='Selected Points'
        )

        self._update_marker()

        self.cid = self.line.figure.canvas.mpl_connect(
            'button_press_event', self.on_click
        )

    def _update_marker(self):
        """Refresh all points from selected_indices."""
        if len(self.selected_indices) == 0:
            self.marker.set_data([], [])
        else:
            x = np.array(self.selected_indices)
            y = self.signal[x]
            self.marker.set_data(x, y)

        self.ax.figure.canvas.draw_idle()
        print(f"-> Selected Indices: {self.selected_indices}")

    def _find_nearest_point(self, x):
        """Find the index of the existing point nearest to x."""
        distances = np.abs(np.array(self.selected_indices) - x)
        return int(np.argmin(distances))

    def on_click(self, event):
        if event.inaxes != self.ax or event.xdata is None:
            return

        new_idx = int(round(event.xdata))
        if not (0 <= new_idx < len(self.signal)):
            return

        # ---------------- Left click: move the nearest point ----------------
        if event.button == 1 and not event.key:
            if len(self.selected_indices) == 0:
                self.selected_indices.append(new_idx)
            else:
                i = self._find_nearest_point(new_idx)
                self.selected_indices[i] = new_idx

        # ---------------- Right click: add a point ----------------
        elif event.button == 3:
            self.selected_indices.append(new_idx)
            self.selected_indices = sorted(set(self.selected_indices))

        # ---------------- Middle click or Shift+left click: delete the nearest point ----------------
        elif event.button == 2 or (event.button == 1 and event.key == 'shift'):
            if len(self.selected_indices) > 0:
                i = self._find_nearest_point(new_idx)
                removed = self.selected_indices.pop(i)
                print(f"-> Removed index: {removed}")

        self._update_marker()


    def get_selected_indices(self):
        """Return all final selected indices."""
        self.line.figure.canvas.mpl_disconnect(self.cid)
        return self.selected_indices
    
def nk_ecg_analysis(signal,fs):
    '''NeuroKit2 ECG Analysis'''
    # Plot the ECG signal
    # plt.plot(signal)
    # plt.show()
    sig_length = signal.shape[-1]

    _, rpeaks = nk.ecg_peaks(signal, sampling_rate=fs,show=False)
    # print(rpeaks)

    # Delineate the ECG signal
    _, waves_peak = nk.ecg_delineate(signal,
                                     rpeaks,
                                     sampling_rate=fs,
                                     method="dwt",  # Can be one of ["dwt", "peak","cwt"]
                                     show=False,
                                     show_type='all')
    # Rename the fields.
    # ECG_R_Onsets-->ECG_Q_Onsets
    waves_peak['ECG_Q_Onsets'] = waves_peak.pop('ECG_R_Onsets')
    # ECG_R_Offsets-->ECG_S_Offsets
    waves_peak['ECG_S_Offsets'] = waves_peak.pop('ECG_R_Offsets')

    # for i in waves_peak.keys():
    #     print(i)
    #     print(waves_peak[i])
    #     print(len(waves_peak[i]))

    # Add the R peak.
    waves_peak['ECG_R_Peaks'] = rpeaks['ECG_R_Peaks']
    # Convert lists in the dictionary to arrays.
    for key, value in waves_peak.items():
        value = np.array(value)
        value = value[~np.isnan(value)]
        value = value.astype(int)
        # Omit values outside the signal range, including negative values.
        value = value[value < sig_length]
        value = value[value > 0]
        waves_peak[key] = value

    # check the result
    # Align the other points by heartbeat according to the R-peak positions.
        # p_onsets, p_peaks, p_offsets, Q_onsets, and Q_peaks precede the R peak.
        # S_peaks, S_offsets, T_onsets, T_peaks, and T_offsets follow the R peak.
    front_points = ['ECG_P_Onsets', 'ECG_P_Peaks', 'ECG_P_Offsets', 'ECG_Q_Onsets', 'ECG_Q_Peaks']
    end_points = ['ECG_S_Peaks', 'ECG_S_Offsets', 'ECG_T_Onsets', 'ECG_T_Peaks', 'ECG_T_Offsets']
    waves_peak_new = {key:np.full(len(waves_peak['ECG_R_Peaks']),-1) for key in waves_peak.keys()}
    incomplete_beats = []
    for i in range(len(waves_peak['ECG_R_Peaks'])):
        waves_peak_new['ECG_R_Peaks'][i] = waves_peak['ECG_R_Peaks'][i]
        for key, value in waves_peak.items():
            # Build a new dictionary keyed like waves_peak and align its values to the R peaks.
            if key in front_points:
                # Points between the previous and current R peaks.
                if i == 0:
                    value_front = 0
                else:
                    value_front = waves_peak['ECG_R_Peaks'][i-1]
                value_end = waves_peak['ECG_R_Peaks'][i]
            elif key in end_points:
                # Points between the current and next R peaks.
                value_front = waves_peak['ECG_R_Peaks'][i]
                if i == len(waves_peak['ECG_R_Peaks'])-1:
                    value_end = sig_length
                else:
                    value_end = waves_peak['ECG_R_Peaks'][i+1]

            if key != 'ECG_R_Peaks':
                # Find values between value_front and value_end.
                value = value[value >= value_front]
                value = value[value < value_end]
                # print(len(waves_peak[key]))
                # print('key',key, 'value',value)
                if len(value) > 0:
                    waves_peak_new[key][i] = value[0]
                else:
                    if i not in incomplete_beats:
                        incomplete_beats.append(i)
    # print(incomplete_beats)
    # # Remove incomplete-beat data.
    # for key, value in waves_peak_new.items():
    #     waves_peak_new[key] = np.delete(value,incomplete_beats)


    # plot
    # plt.figure(figsize=(10, 5))
    # plt.plot(signal)
    # # plt.scatter(rpeaks['ECG_R_Peaks'], signal[rpeaks['ECG_R_Peaks']], color='red')
    # for key, value in waves_peak.items():
    #     if key == 'ECG_R_Peaks':
    #         plt.scatter(value, signal[value], color='red',label=key)
    #     # Convert lists to arrays and omit NaN values.
    #     else:
    #         plt.scatter(value, signal[value],label=key)
    # plt.legend()
    # plt.show()

    # print(waves_peak_new)
    # Heart-rate variability.
    # ecg_hrv = nk.ecg_hrv(signal, rpeaks, sampling_rate=fs)
    return waves_peak_new

def median_filter_multichannel(signal, fs):
    """
    Apply median baseline-correction filtering to multiple leads.
    
    The input signal is assumed to have shape (number of leads, number of samples).
    
    :param signal: Input signal with shape (M, N), where M is the number of leads and N is the number of samples.
    :param fs: Sampling frequency.
    :return: Filtered signal with the same shape as the input.
    """
    
    # Ensure the signal is 2D; otherwise raise an error or convert it to (1, N).
    if signal.ndim == 1:
        # Convert a 1D input to (1, N) for consistent processing.
        signal = signal[np.newaxis, :]
    
    num_leads, num_samples = signal.shape
    
    # Compute the filter size from the sampling frequency to remove low-frequency drift near periods below 0.8 s.
    kernel_size = int(fs * 0.8)
    # kernel_size = 5
    
    
    # Keep kernel_size odd to simplify padding.
    if kernel_size % 2 == 0:
        kernel_size += 1
        
    pad_width = (kernel_size - 1) // 2
    
    # Initialize the result matrix.
    filtered_signal = np.zeros_like(signal, dtype=signal.dtype)
    
    # Iterate over each lead (row).
    for lead_idx in range(num_leads):
        single_lead = signal[lead_idx, :]
        
        # 1. Pad both ends of the current lead.
        # Use edge mode to reduce boundary effects.
        padded_lead = np.pad(single_lead, pad_width, mode='edge')
        
        # 2. Apply median-based baseline correction.
        # Iterate over each sample.
        for i in range(num_samples):
            # Extract the sliding window.
            window = padded_lead[i:i + kernel_size]
            
            # Subtract the local median (the baseline trend) from the current value.
            filtered_signal[lead_idx, i] = single_lead[i] - np.median(window)
            
    return filtered_signal

def list_subfolders(path):
    """obtain the list of subfolders in a given path"""
    return [f for f in os.listdir(path)
            if os.path.isdir(os.path.join(path, f))]
    
# ----------------CUT ECG TO QRST -----------
def cut_ecg_batch():
    '''extract QRS-T segments from ECG signal batch'''
    # ----------------------------------------------------------------
    # Ensure list_subfolders, median_filter, and nk_ecg_analysis are defined.
    # ----------------------------------------------------------------
    # data_dir = r"/path/to/dataset/paired_ECG_heartbeats"
    # ourput_dir = r"/path/to/output/paired_ECG_heartbeats_QRST"
    data_dir = r"/path/to/dataset/ECG_series_paired_heartbeats"
    ourput_dir = r"/path/to/output/ECG_series_paired_heartbeats_QRST"
    os.makedirs(ourput_dir, exist_ok=True)
    ecgs = os.listdir(data_dir)
    fs = 500 

    for ecg in ecgs:
        subject_id = ecg.split('.npy')[0]

        # try:
        # Read the data.
        # load .npy data
        data = np.load(os.path.join(data_dir, ecg)) 
        data_filtered = median_filter_multichannel(data, fs)
        filtered_signal = data_filtered[1, :]
        signal_length = len(filtered_signal)

        # # Preprocess the signal.
        # filtered_signal = median_filter(signal_original, fs)
        
        # --- 1. Automatic localization (use the tiling workaround to keep NK2 working). ---
        
        # Tile the signal three times to reduce computation and simplify index handling.
        ecg_data_tiled = np.tile(filtered_signal, 10) 
        ecg_points = nk_ecg_analysis(ecg_data_tiled, fs)
        Q_peaks_all = ecg_points.get('ECG_Q_Peaks', np.array([]))
        
        # Find the Q peak most likely to correspond to the central beat of the original signal.
        Q_peaks_candidate = Q_peaks_all[
            (Q_peaks_all >= signal_length) & (Q_peaks_all < signal_length * 2)
        ]
        
        if Q_peaks_candidate.size > 0:
            # Find the point nearest the first-beat center (L/2) and convert it to an original-signal index.
            # Assume the first complete beat lies between indices L and 2L in the tiled data.
            target_center = signal_length + signal_length // 2
            best_peak_tiled_idx = Q_peaks_candidate[np.argmin(np.abs(Q_peaks_candidate - target_center))]
            
            # Convert back to an original-signal index in [0, L-1].
            initial_Q_peak = best_peak_tiled_idx - signal_length 
            print(f"Auto-detected Q-peak at index: {initial_Q_peak}")

        else:
            initial_Q_peak = None
            print("Auto-detection FAILED (No Q-peak found in the expected range). Entering manual selection.")

        # --- 2. Interactive correction/definition. ---

        # Enter interactive mode if automatic detection fails, returns multiple points, or needs inspection.
        # if initial_Q_peak is None or Q_peaks_candidate.size != 1:
        #     manual_selection_mode = True
        # else:
        #     # Optionally ask whether correction is needed; enter directly here for simplicity.
        #     manual_selection_mode = False 
        manual_selection_mode  = True
        
        final_Q_peak = initial_Q_peak
        
        if manual_selection_mode or final_Q_peak is None:
            fig, ax = plt.subplots(figsize=(12, 6))
            
            # Plot both raw and filtered signals to aid baseline assessment.
            ax.plot(data[1,:], label='Original Signal', color='gray', alpha=0.5)
            
            # Start the interactive selector.
            selector = InteractiveSelector(
                ax, 
                filtered_signal, 
                initial_peak_idx=final_Q_peak,
                title=f'Subject {subject_id} - Manual QRS Onset Selection'
            )
            
            plt.legend()
            plt.show(block=True) # Block until the window closes.
            
            final_Q_peak = selector.get_selected_indices()[0]
        
        # --- 3. Output the results. ---
        
        if final_Q_peak is not None and final_Q_peak != -1:
            print(f"Final Q-peak for {subject_id}: {final_Q_peak}")
            # Save final_Q_peak here, for example in a dictionary or CSV file.
            data_final = data_filtered[:, final_Q_peak:]  # Extract QRST.
            
            # Resample to 1000 Hz.
            num_samples_original = data_final.shape[1]
            num_samples_resampled = int(num_samples_original * (1000 / fs))
            data_final = scipy.signal.resample(data_final, num_samples_resampled, axis=1)
            # Save the result.
            np.save(os.path.join(ourput_dir, f"{subject_id}_QRST.npy"), data_final)
        else:
            print(f"Warning: Q-peak not selected/found for {subject_id}")
        
        # Optional final visual verification.
        # plt.figure(figsize=(10, 5))
        # plt.title(f'Subject {subject_id} - Final Result')
        # plt.plot(signal_original, label='Original')
        # plt.plot(filtered_signal, label='Filtered')
        # if final_Q_peak is not None and final_Q_peak != -1:
        #     plt.scatter(final_Q_peak, filtered_signal[final_Q_peak], color='red', zorder=5, label='Final Q Peak')
        # plt.legend()
        # plt.show()

        # except FileNotFoundError:
        #     print(f"Error: File not found for {subject_id}")
        # except Exception as e:
        #     print(f"An unexpected error occurred for {subject_id}: {e}")
        
def cut_ecg_batch_NUH():
    data_dir = r"/path/to/dataset/ECG_series_paired_heartbeats"
    ourput_dir = r"/path/to/output/ECG_series_paired_heartbeats_QRST"
    os.makedirs(ourput_dir, exist_ok=True)
    patient_dirs = list_subfolders(data_dir)
    for patient in patient_dirs:
        os.makedirs(os.path.join(ourput_dir, patient), exist_ok=True)
        ecgs = os.listdir(os.path.join(data_dir,patient))
        fs = 500 

        for ecg in ecgs:
            subject_id = ecg.split('.npy')[0]

            # try:
            # Read the data.
            # load .npy data
            data = np.load(os.path.join(data_dir, patient, ecg)) 
            data_filtered = median_filter_multichannel(data, fs)
            filtered_signal = data_filtered[1, :]
            signal_length = len(filtered_signal)

            # # Preprocess the signal.
            # filtered_signal = median_filter(signal_original, fs)
            
            # --- 1. Automatic localization (use the tiling workaround to keep NK2 working). ---
            
            # Tile the signal three times to reduce computation and simplify index handling.
            ecg_data_tiled = np.tile(filtered_signal, 10) 
            ecg_points = nk_ecg_analysis(ecg_data_tiled, fs)
            Q_peaks_all = ecg_points.get('ECG_Q_Peaks', np.array([]))
            
            # Find the Q peak most likely to correspond to the central beat of the original signal.
            Q_peaks_candidate = Q_peaks_all[
                (Q_peaks_all >= signal_length) & (Q_peaks_all < signal_length * 2)
            ]
            
            if Q_peaks_candidate.size > 0:
                # Find the point nearest the first-beat center (L/2) and convert it to an original-signal index.
                # Assume the first complete beat lies between indices L and 2L in the tiled data.
                target_center = signal_length + signal_length // 2
                best_peak_tiled_idx = Q_peaks_candidate[np.argmin(np.abs(Q_peaks_candidate - target_center))]
                
                # Convert back to an original-signal index in [0, L-1].
                initial_Q_peak = best_peak_tiled_idx - signal_length 
                print(f"Auto-detected Q-peak at index: {initial_Q_peak}")

            else:
                initial_Q_peak = None
                print("Auto-detection FAILED (No Q-peak found in the expected range). Entering manual selection.")

            # --- 2. Interactive correction/definition. ---

            # Enter interactive mode if automatic detection fails, returns multiple points, or needs inspection.
            # if initial_Q_peak is None or Q_peaks_candidate.size != 1:
            #     manual_selection_mode = True
            # else:
            #     # Optionally ask whether correction is needed; enter directly here for simplicity.
            #     manual_selection_mode = False 
            manual_selection_mode  = True
            
            final_Q_peak = initial_Q_peak
            
            if manual_selection_mode or final_Q_peak is None:
                fig, ax = plt.subplots(figsize=(12, 6))
                
                # Plot both raw and filtered signals to aid baseline assessment.
                ax.plot(data[1,:], label='Original Signal', color='gray', alpha=0.5)
                
                # Start the interactive selector.
                selector = InteractiveSelector(
                    ax, 
                    filtered_signal, 
                    initial_peak_idx=final_Q_peak,
                    title=f'Subject {subject_id} - Manual QRS Onset Selection'
                )
                
                plt.legend()
                plt.show(block=True) # Block until the window closes.
                
                final_Q_peak = selector.get_selected_indices()[0]
            
            # --- 3. Output the results. ---
            
            if final_Q_peak is not None and final_Q_peak != -1:
                print(f"Final Q-peak for {subject_id}: {final_Q_peak}")
                # Save final_Q_peak here, for example in a dictionary or CSV file.
                data_final = data_filtered[:, final_Q_peak:]  # Extract QRST.
                
                # Resample to 1000 Hz.
                num_samples_original = data_final.shape[1]
                num_samples_resampled = int(num_samples_original * (1000 / fs))
                data_final = scipy.signal.resample(data_final, num_samples_resampled, axis=1)
                # Save the result.
                np.save(os.path.join(ourput_dir,patient, f"{subject_id}_QRST.npy"), data_final)
            else:
                print(f"Warning: Q-peak not selected/found for {subject_id}")
            
            # Optional final visual verification.
            # plt.figure(figsize=(10, 5))
            # plt.title(f'Subject {subject_id} - Final Result')
            # plt.plot(signal_original, label='Original')
            # plt.plot(filtered_signal, label='Filtered')
            # if final_Q_peak is not None and final_Q_peak != -1:
            #     plt.scatter(final_Q_peak, filtered_signal[final_Q_peak], color='red', zorder=5, label='Final Q Peak')
            # plt.legend()
            # plt.show()

            # except FileNotFoundError:
            #     print(f"Error: File not found for {subject_id}")
            # except Exception as e:
            #     print(f"An unexpected error occurred for {subject_id}: {e}")
def plot_npy_signal(path):

    data = np.load(path)
    
    fig, axs = plt.subplots(12, 1, figsize=(3, 10))
    lead_names = ['I', 'II', 'III', 'aVR', 'aVL', 'aVF', 'V1', 'V2', 'V3', 'V4', 'V5', 'V6']
    for i in range(12):
        axs[i].plot(data[i], color='tab:blue', lw=1)
        axs[i].set_ylabel(lead_names[i], rotation=0, labelpad=20, fontsize=8, fontweight='bold')
        axs[i].set_yticks([])
        if i < 11: axs[i].set_xticks([])
        axs[i].grid(True, alpha=0.2, ls=':')
    output_path = path.replace('.npy', '.png')
    # plt.savefig(output_path, dpi=300, bbox_inches='tight')
    print(data.shape)
    plt.show()
    
def plot_npy_signal_batch():
    # path = r"/path/to/dataset/paired_ECG_heartbeats_QRST/sample_QRST.npy"
    data_dir = r"/path/to/dataset/paired_ECG_heartbeats_QRST"
    out_csv = os.path.join(data_dir, "npy_summary.csv")
    fs = 1000  # happen in cut_ecg_batch()
    records = []
    for fname in sorted(os.listdir(data_dir)):
        if not fname.endswith(".npy"):
            continue
        fpath = os.path.join(data_dir, fname)
        data = np.load(fpath)
        records.append({
            "filename": fname,
            "fs": fs,
            "shape": str(data.shape),
        })
        # plot
        # plot_npy_signal(fpath)
    df = pd.DataFrame(records)
    df.to_csv(out_csv, index=False)
    print(f"Saved to {out_csv}")

            

# ------------------------------ECG end------------------------------- #

def check_ecg_digitization_adaptive(digitized_dir, raw_img_dir, output_csv="check_ecg_results.csv"):
    '''check and visualize ECG digitization results with adaptive folder naming'''
    # 1. Scan and group the files.
    output_csv = os.path.join(digitized_dir, output_csv)
    all_subfolders = [f for f in os.listdir(digitized_dir) if os.path.isdir(os.path.join(digitized_dir, f))]
    groups = defaultdict(list)
    for f in all_subfolders:
        parts = f.split('_')
        base_id = f"{parts[0]}_{parts[1]}" if len(parts) >= 2 else f
        groups[base_id].append(f)

    # --- Resume-from-checkpoint logic. ---
    checked_ids = set()
    results = []
    if os.path.exists(output_csv):
        try:
            existing_df = pd.read_csv(output_csv)
            # Assume the first CSV column is 'id'.
            if 'id' in existing_df.columns:
                checked_ids = set(existing_df['id'].astype(str).tolist())
                results = existing_df.to_dict('records') # Load existing records.
                print(f"检测到已存在的进度：已跳过 {len(checked_ids)} 个已校验项目。")
        except Exception as e:
            print(f"读取旧 CSV 失败，将重新开始: {e}")

    # 2. Iterate over each group.
    for base_id, folder_list in groups.items():
        # Select the target folder.
        target_folder = None
        seg_1 = next((f for f in folder_list if f.endswith('_1')), None)
        if seg_1:
            target_folder = seg_1
        elif len(folder_list) == 1:
            target_folder = folder_list[0]
        else:
            continue

        # --- Check whether this item has already been validated. ---
        if target_folder in checked_ids:
            continue 

        # 3. Match the BMP image.
        img_path = os.path.join(raw_img_dir, f"{target_folder}.bmp")
        if not os.path.exists(img_path):
            img_path = os.path.join(raw_img_dir, f"{base_id}.bmp")
        
        if not os.path.exists(img_path):
            print(f"跳过 {target_folder}: 找不到对应的图片")
            continue

        # 4. Read the CSV and plot it.
        csv_files = glob.glob(os.path.join(digitized_dir, target_folder, "*.csv"))
        if not csv_files: continue

        try:
            df = pd.read_csv(csv_files[0], header=None)
            data = df.values.T if df.shape[1] == 12 else df.values
            if data.shape[0] != 12: continue

            # --- Visualization and interaction. ---
            lead_names = ['I', 'II', 'III', 'aVR', 'aVL', 'aVF', 'V1', 'V2', 'V3', 'V4', 'V5', 'V6']
            fig = plt.figure(figsize=(16, 9))
            gs = fig.add_gridspec(12, 2, width_ratios=[2, 1])

            # Left: original image.
            ax_img = fig.add_subplot(gs[:, 0])
            with Image.open(img_path) as img:
                ax_img.imshow(img)
            ax_img.set_title(f"Original Image: {target_folder}", fontsize=10)
            ax_img.axis('off')

            # Right: signal.
            for i in range(12):
                ax_lead = fig.add_subplot(gs[i, 1])
                ax_lead.plot(data[i, :], color='tab:red', lw=0.8)
                ax_lead.set_ylabel(lead_names[i], rotation=0, labelpad=20, fontsize=8, fontweight='bold')
                ax_lead.set_yticks([])
                if i < 11: ax_lead.set_xticks([])
                ax_lead.grid(True, alpha=0.2, ls=':')

            plt.suptitle(f"Digitization Verification: {target_folder}", fontsize=14, y=0.96)
            
            # --- Interactive-button logic. ---
            current_status = {"id": target_folder, "status": None}

            def on_click_pass(event):
                current_status["status"] = 1
                plt.close(fig)

            def on_click_fail(event):
                current_status["status"] = 0
                plt.close(fig)

            # Lay out the buttons.
            ax_pass = plt.axes([0.43, 0.02, 0.06, 0.04])
            ax_fail = plt.axes([0.51, 0.02, 0.06, 0.04])
            btn_pass = Button(ax_pass, 'Pass (1)', color='limegreen', hovercolor='green')
            btn_fail = Button(ax_fail, 'Fail (0)', color='tomato', hovercolor='red')
            
            btn_pass.on_clicked(on_click_pass)
            btn_fail.on_clicked(on_click_fail)

            plt.show() # Block while waiting for the user to click.

            # 5. Save the result.
            if current_status["status"] is not None:
                results.append(current_status)
                # Save immediately after each click, overwriting the old file to update progress.
                pd.DataFrame(results).to_csv(output_csv, index=False, encoding='utf-8')
                print(f"已记录: {target_folder} -> {current_status['status']}")
            else:
                # If the window is closed without a button click, do not record a result so it appears again next time.
                print(f"未对 {target_folder} 进行标记，已跳过。")

        except Exception as e:
            print(f"处理 {target_folder} 出错: {e}")

    print("-" * 30)
    print(f"校验任务结束。最终结果保存在: {output_csv}")

def plot_ecg_image(target_folder, digitized_dir, raw_img_dir):
    # 3. Match the BMP image.
    img_path = os.path.join(raw_img_dir, f"{target_folder}.bmp")

    # 4. Read the CSV and plot it.
    csv_files = glob.glob(os.path.join(digitized_dir, target_folder, "*.csv"))

    df = pd.read_csv(csv_files[0], header=None)
    data = df.values.T if df.shape[1] == 12 else df.values

    # --- Visualization and interaction. ---
    lead_names = ['I', 'II', 'III', 'aVR', 'aVL', 'aVF', 'V1', 'V2', 'V3', 'V4', 'V5', 'V6']
    fig = plt.figure(figsize=(16, 9))
    gs = fig.add_gridspec(12, 2, width_ratios=[2, 1])

    # Left: original image.
    ax_img = fig.add_subplot(gs[:, 0])
    with Image.open(img_path) as img:
        ax_img.imshow(img)
    ax_img.set_title(f"Original Image: {target_folder}", fontsize=10)
    ax_img.axis('off')

    # Right: signal.
    for i in range(12):
        ax_lead = fig.add_subplot(gs[i, 1])
        ax_lead.plot(data[i, :], color='tab:red', lw=0.8)
        ax_lead.set_ylabel(lead_names[i], rotation=0, labelpad=20, fontsize=8, fontweight='bold')
        ax_lead.set_yticks([])
        if i < 11: ax_lead.set_xticks([])
        ax_lead.grid(True, alpha=0.2, ls=':')

    plt.suptitle(f"Digitization Verification: {target_folder}", fontsize=14, y=0.96)
    
    # --- Interactive-button logic. ---
    current_status = {"id": target_folder, "status": None}

    def on_click_pass(event):
        current_status["status"] = 1
        plt.close(fig)

    def on_click_fail(event):
        current_status["status"] = 0
        plt.close(fig)

    plt.show() # Block while waiting for the user to click.

def obtain_median_heartbeat(ECG_path, fs = 409.6, data_length = 4096):
    '''Core script: obtain median heartbeat from ECG signal'''
    print(ECG_path)
    df = pd.read_csv(ECG_path, header=None)
    data = df.values.T if df.shape[1] == 12 else df.values
    print(data.shape)  # (12, 4096)
    segments = {
        "I": (0, data_length/4), "II": (0, data_length), "III": (0, data_length/4),
        "aVR": (data_length/4, data_length/4*2), "aVL": (data_length/4, data_length/4*2), "aVF": (data_length/4, data_length/4*2),
        "V1": (data_length/4*2, data_length/4*3), "V2": (data_length/4*2, data_length/4*3), "V3": (data_length/4*2, data_length/4*3),
        "V4": (data_length/4*3, data_length), "V5": (data_length/4*3, data_length), "V6": (data_length/4*3, data_length)
    }

    lead_names = ["I", "II", "III", "aVR", "aVL", "aVF", "V1", "V2", "V3", "V4", "V5", "V6"]
    
    # signals, info = nk.ecg_process(data[1, : ], sampling_rate=409.6) 
    # fs = 409.6
    _, rpeaks = nk.ecg_peaks(data[1, : ], sampling_rate=fs,show=False)
    r_peaks = rpeaks['ECG_R_Peaks']
    
    # Manually correct R-wave positions.
    fig, ax = plt.subplots(figsize=(12, 4))
    
    # Plot both raw and filtered signals to aid baseline assessment.
    ax.plot(data[1,:], label='Original Signal', color='gray', alpha=0.5)
    
    # Start the interactive selector.
    selector = InteractiveSelector(
        ax, 
        data[1, : ], 
        initial_peak_idx=r_peaks,
        title=f'{ECG_path}'
    )
    
    plt.legend()
    plt.show(block=True) # Block until the window closes.
    
    r_peaks = selector.get_selected_indices()
    r_peaks = np.array(r_peaks)
    print(r_peaks)
    
    
    # Keep only R waves within the valid time range of this lead.
    start, end = segments['II']
    diff = np.diff(r_peaks)
    average_rr = np.median(diff)
    # plot II lead with r peaks
    # plt.plot(data[1, : ], color='tab:blue', lw=1)
    # plt.plot(r_peaks, data[1, r_peaks], 'ro')
    
    mean_beats = []
    for idx, name in enumerate(lead_names):
        start, end = segments[name]
        # print(start, end)
        # if name in ["I", "II", "III"]:
        #     dominate_idx = 1  # Lead II is the dominant lead.
        # elif name in ["aVR", "aVL", "aVF"]:
        #     dominate_idx = 5  # Lead aVF is the dominant lead.
        # elif name in ["V1", "V2", "V3"]:
        #     dominate_idx = 8  # Lead V3 is the dominant lead.
        # else:
        #     dominate_idx = 11 # Lead V6 is the dominant lead.
            
        # signals, info = nk.ecg_process(data[dominate_idx, : ], sampling_rate=409.6, quality=False)
        # r_peaks = info['ECG_R_Peaks']
        
        # Keep only R waves within the valid time range of this lead.
        current_r_peaks = r_peaks[(r_peaks >= start) & (r_peaks <= end)]
        print(current_r_peaks)
        

        # plt.figure()
        # plt.plot(data[idx, : ], color='tab:blue', lw=1)
        # plt.plot(cloest_R, data[idx, cloest_R], 'ro')
        # plt.show()
        
        # # Extract beats and compute the median beat.
        # Method 1: extract beats directly with NeuroKit2.
        # epochs = nk.ecg_segment(data[idx, : ], r_peaks=current_r_peaks, sampling_rate=409.6,show=True)
        # plt.show()
        # Method 2: group R points, then extract fixed-length beats using the lead-II heart rate.
        # epochs = {}
        # for i in current_r_peaks:
        #     if i - int(0.3 * average_rr) < start or i + int(0.7 * average_rr) > end:
        #         continue
        #     epoch = data[idx, i - int(0.4 * average_rr): i + int(0.6 * average_rr)]
        #     if 'epochs' not in locals():
        #         epochs = {}
        #     epochs[str(i)] = {"Signal": epoch}


        # # Convert dictionary-form epochs to their mean.
        # beats = [epochs[str(k)]["Signal"] for k in epochs.keys()]
        # mean_beats.append(np.median(beats, axis=0))
        # Method 3: use only the R wave nearest the center.
        cloest_R = current_r_peaks[np.argmin(np.abs(current_r_peaks - (start + end) / 2))]
        range_start = cloest_R - int(0.4 * average_rr)
        range_end = cloest_R + int(0.6 * average_rr)
        heartbeat = data[idx, range_start: range_end]
        if heartbeat.shape[0] < int(average_rr):
            heartbeat = np.pad(heartbeat, (0, int(average_rr) - heartbeat.shape[0]), mode='constant')
        mean_beats.append(heartbeat)
        
        # FIXME: Extract QRST.
        # cloest_Q_onset = Q_onsets[(Q_onsets >= range_start) & (Q_onsets <= range_end)]
        # heartbeat_QRST = data[idx,cloest_Q_onset:range_end]
        # QRST_bests.append(heartbeat_QRST)
        
        # plt.plot(mean_beats[-1], color='tab:orange', lw=1)
        # plt.show()
    # plot
    fig, axs = plt.subplots(12, 1, figsize=(3, 7))
    lead_names = ['I', 'II', 'III', 'aVR', 'aVL', 'aVF', 'V1', 'V2', 'V3', 'V4', 'V5', 'V6']
    for i in range(12):
        axs[i].plot(mean_beats[i], color='tab:blue', lw=1)
        axs[i].set_ylabel(lead_names[i], rotation=0, labelpad=20, fontsize=8, fontweight='bold')
        axs[i].set_yticks([])
        if i < 11: axs[i].set_xticks([])
        axs[i].grid(True, alpha=0.2, ls=':')
    
    plt.show()
    
    # fig, axs = plt.subplots(12, 1, figsize=(10, 15))
    # lead_names = ['I', 'II', 'III', 'aVR', 'aVL', 'aVF', 'V1', 'V2', 'V3', 'V4', 'V5', 'V6']
    # for i in range(12):
    #     axs[i].plot(QRST_bests[i], color='tab:blue', lw=1)
    #     axs[i].set_ylabel(lead_names[i], rotation=0, labelpad=20, fontsize=8, fontweight='bold')
    #     axs[i].set_yticks([])
    #     if i < 11: axs[i].set_xticks([])
    #     axs[i].grid(True, alpha=0.2, ls=':')
    
    # plt.show()
    return mean_beats
    
def obtain_median_heartbeat_batch():
    paired_csv_path = "/path/to/dataset/paired_ecg_mri.csv"
    df_paired = pd.read_csv(paired_csv_path)
    output_dir = "/path/to/output/paired_ECG_heartbeats"
    
    digitized_dir = "/path/to/dataset/data_ECG_digitalized"
    raw_img_dir = "/path/to/dataset/data_ECG"
    
    df_paired = df_paired[df_paired['need_run']==1]
    os.makedirs(output_dir, exist_ok=True)
    for index, row in df_paired.iterrows():
        ecg_folder_name = row['ECG_Folder']
        ECG_dir = os.path.join(digitized_dir, ecg_folder_name)
        # Read the first CSV file in the directory.
        ECG_path = glob.glob(os.path.join(ECG_dir, "*.csv"))[0]
        # Read the original image.
        plot_ecg_image(target_folder=ecg_folder_name, digitized_dir=digitized_dir, raw_img_dir=raw_img_dir)
        mean_beats = obtain_median_heartbeat(ECG_path)
        np_mean_beats = np.array(mean_beats)
        print(np_mean_beats.shape)  # (12, N)
        np.save(os.path.join(output_dir, f"{ecg_folder_name}.npy"), np_mean_beats)
        
def obtain_median_heartbeat_batch_NUH():
    paired_csv_path = "/path/to/dataset/paired_ecg_mri.csv"
    df_paired = pd.read_csv(paired_csv_path)
    output_dir = "/path/to/output/ECG_series_paired_heartbeats"
    
    digitized_dir = "/path/to/dataset/ECG_series_digitalized_full"
    
    # df_paired = df_paired[df_paired['need_run']==1]
    os.makedirs(output_dir, exist_ok=True)
    for index, row in df_paired.iterrows():
        item_name = row['ECG_Folder'].split('\\')[0]
        print('item_name:', item_name)
        os.makedirs(os.path.join(output_dir, item_name), exist_ok=True)
        ecg_folder_name = row['ECG_Folder'].split('.pdf')[0]
        if os.path.exists(os.path.join(output_dir, f"{ecg_folder_name}.npy")):
            print(f"File already exists, skipping: {row['ECG_Folder']}")
            continue
        ECG_path = os.path.join(digitized_dir, ecg_folder_name + ".csv")
        # Read the original image.
        mean_beats = obtain_median_heartbeat(ECG_path, fs = 500, data_length=5000)
        np_mean_beats = np.array(mean_beats)
        print(np_mean_beats.shape)  # (12, N)
        np.save(os.path.join(output_dir, f"{ecg_folder_name}.npy"), np_mean_beats)
        
def load_ecg_npy_8leaads(ECG_path):
    '''load ECG npy file'''
    data = np.load(ECG_path)
    # only consider I II, V1-V6 leads
    data = data[[0,1,6,7,8,9,10,11], :]
    print(f"Loaded ECG shape: {data.shape} from {ECG_path}")
    return data

def VTK_merged_ED_scar():
    '''copy 3dscar_gd.vtk files to a new folder'''
    data_dir = "/path/to/dataset/VTK_Merged_with_OPT_Process"
    output_dir = "/path/to/output/VTK_Merged_ED"
    os.makedirs(output_dir, exist_ok=True)
    
    # Copy files containing 3dscar_gd.vtk to output_dir.
    list = []
    for folder in os.listdir(data_dir):
        folder_path = os.path.join(data_dir, folder)
        if not os.path.isdir(folder_path):
            continue
            
        vtk_files = glob.glob(os.path.join(folder_path, "*3dscar_gd.vtk"))
        if vtk_files:
            for vtk_file in vtk_files:
                base_name = os.path.basename(vtk_file)
                output_path = os.path.join(output_dir, base_name)
                
                # Replace os.rename with shutil.copy2.
                shutil.copy2(vtk_file, output_path)
                print(f"Copied: {vtk_file} -> {output_path}")
                
                list.append(base_name)
    # save list to csv, column name is filename
    # add index column, start from 1,
    # add biggestVentRV,need_run,valid column
    df = pd.DataFrame(list, columns=["filename"])
    df.index += 1
    df["biggestVentRV"] = 0
    df["need_run"] = 1
    df["valid"] = 0
    
    df.to_csv(os.path.join(output_dir, "mesh_biggestVentRV.csv"), index_label="index")
    
    

def paired_ecg_mri(CONFIG, paired_csv_path):
    
    '''(Fudan)Return paired ECG and MRI data based on metadata and validity checks'''
    
    
    meta_path = CONFIG["meta"]
    ecg_dir = CONFIG["ecg"]
    mri_dir = CONFIG["mri"]
    check_csv_path = CONFIG["csv"]
    # 1. Load invalid ECG list from CSV
    check_df = pd.read_csv(check_csv_path)
    invalid_ecg_folders = set(check_df[check_df.iloc[:, 1] == 0].iloc[:, 0].astype(str).tolist())

    # 2. Map ECG folders: { patient_id: [(folder_name, date_obj), ...] }
    pattern = re.compile(r"^(patient\d+)_(\d{8})(_1)?$")
    ecg_mapping = {}
    for folder in os.listdir(ecg_dir):
        match = pattern.match(folder)
        if match and folder not in invalid_ecg_folders:
            p_id = match.group(1)
            try:
                date_obj = datetime.strptime(match.group(2), "%Y%m%d")
                if p_id not in ecg_mapping:
                    ecg_mapping[p_id] = []
                ecg_mapping[p_id].append((folder, date_obj))
            except ValueError:
                continue
            
    # 2. Map MRI folders: { patient_id: [(folder_name, date_obj), ...] }
    mri_mapping = {}
    pattern = re.compile(r"^(patient\d+)_(\d{8})")
    for folder in os.listdir(mri_dir):
        match = pattern.match(folder)
        if match:
            p_id = match.group(1)
            try:
                date_obj = datetime.strptime(match.group(2), "%Y%m%d")
                if p_id not in mri_mapping:
                    mri_mapping[p_id] = []
                mri_mapping[p_id].append((folder, date_obj))
            except ValueError:
                continue
    # print(mri_mapping)

    # 3. Process Meta Excel and Pair Data

    wb = openpyxl.load_workbook(meta_path)
    sheet = wb.active
    header = [cell.value for cell in sheet[1]]

    idx_id = header.index("Patient_ID") + 1
    idx_mesh = header.index("Mesh_valid") + 1
    # idx_mri_date = header.index("MRI_date") + 1
    idx_ecg_valid = header.index("ECG_valid") + 1

    print(f"{'Patient_ID':<15} | {'MRI Date':<10} | {'Match Count':<12} | {'Best Match'}")
    print("-" * 75)

    for row_idx in range(2, sheet.max_row + 1):
        p_id_val = sheet.cell(row=row_idx, column=idx_id).value
        mesh_val_folder_names = [item[0] for item in mri_mapping.get(p_id_val, [])]
        ecg_val_folder_names = [item[0] for item in ecg_mapping.get(p_id_val, [])]

        # Update Excel Cell
        sheet.cell(row=row_idx, column=idx_mesh).value = ",".join(mesh_val_folder_names)
        sheet.cell(row=row_idx, column=idx_ecg_valid).value = ",".join(ecg_val_folder_names)
    wb.save(meta_path)
    print(f"\n Processing complete. Metadata saved to: {meta_path}")
        
    # TODO:Pair MRI and ECG (one-to-one), ensure every MRI comes with one ECG, the datatime bwteen them is the closest, and <30 days
    # save to a new csv
    # 4. Pair each MRI with its nearest ECG when the interval is less than 30 days.
    pairs = []
    pairing_results = [] # Used for CSV export.

    for p_id, mris in mri_mapping.items():
        if p_id not in ecg_mapping:
            continue
        
        available_ecgs = ecg_mapping[p_id]
        
        for mri_folder, mri_date in mris:
            best_ecg = None
            min_diff = 31 # Initialize to 31 days.

            for ecg_folder, ecg_date in available_ecgs:
                diff = abs((mri_date - ecg_date).days)
                if min_diff is None:
                    min_diff = diff
                    best_ecg = ecg_folder
                elif diff < min_diff:
                    min_diff = diff
                    best_ecg = ecg_folder
            
            if best_ecg:
                pairs.append((mri_folder, best_ecg))
                pairing_results.append({
                    "Patient_ID": p_id,
                    "Mesh_Folder": mri_folder,
                    "ECG_Folder": best_ecg,
                    "Days_Diff": min_diff
                })

    # 5. Save the pairing results to a new CSV file.
    pd.DataFrame(pairing_results).to_csv(paired_csv_path, index=False)
    print(f"Pairing complete. {len(pairs)} pairs found. Saved to: {paired_csv_path}")


def get_header_datetime_obj(pdf_path, header_height=50):
    # 1. Read text from the top region of the PDF.
    doc = pymupdf.open(pdf_path)
    page = doc[0]
    # header_rect = pymupdf.Rect(0, 0, page.rect.width, header_height)
    header_text = page.get_text("text").strip()
    # print(f"Extracted header text from '{pdf_path}': '{header_text}'")
    doc.close()
    # Allow a more flexible time format in case some files omit seconds (\d{2}:\d{2}(:\d{2})?).
    pattern = r"\d{2}-[A-Za-z]{3}-\d{4}\s+\d{2}:\d{2}(?::\d{2})?"
    safe_text = header_text.split("When compared with")[0]
    match = re.search(pattern, safe_text)

    if match:
        extracted_date = match.group(0)
        print(f"提取到的独立日期为: '{extracted_date}'")
    else:
        print("未在截断后的文本中找到匹配的日期。")
        
    return datetime.strptime(extracted_date, "%d-%b-%Y %H:%M:%S") if len(extracted_date.split(':')) == 3 else datetime.strptime(extracted_date, "%d-%b-%Y %H:%M")

def paired_ecg_mri_NUH(CONFIG):
    
    '''(NUH)Return paired ECG and MRI data'''
    ecg_dir = CONFIG["ecg"]
    mri_dir = CONFIG["mri"]
    metadata_path = CONFIG["meta"]
    paired_csv_path = CONFIG["csv"]
    # 1. Map ECG folders: { patient_id: [(path_name, date_obj), ...] }
    pattern = re.compile(r"N(\w+)-([A-Z]+)")
    ecg_mapping = {}
    ECG_states = {}
    for folder in os.listdir(ecg_dir):
        # folder = "N001-SNP"
        match = pattern.match(folder)
        if match:
            p_id = match.group(1)
            ECG_states[p_id] = match.group(2)
            for file in os.listdir(os.path.join(ecg_dir, folder)):
                if file.endswith(".pdf"):
                    pdf_path = os.path.join(ecg_dir, folder, file)
                    
                    try:
                        # TODO: Find the date in the PDF, formatted as 07-Jul-2015 11:13:23.
                        dt = get_header_datetime_obj(pdf_path)
                        date_obj = dt
                        if p_id not in ecg_mapping:
                            ecg_mapping[p_id] = []
                        ecg_mapping[p_id].append((os.path.join(folder,file), date_obj))
                    except ValueError:
                        continue
    print(ecg_mapping)
            
    # 2. Map MRI folders: { patient_id: [(path_name, date_obj), ...] }
    mri_mapping = {}
    pattern = re.compile(r"IMMC(\d{3})_(\d{8})")
    for folder in os.listdir(mri_dir):
        match = pattern.match(folder)
        
        if match:
            p_id = match.group(1)
            try:
                date_obj = datetime.strptime(match.group(2), "%Y%m%d")
                if p_id not in mri_mapping:
                    mri_mapping[p_id] = []
                mri_mapping[p_id].append((folder, date_obj))
            except ValueError:
                continue
    print(mri_mapping)

    # 3. Process Meta Excel and Pair Data
    # columns: Patient_ID,MRI_date,ECG_date
     # Take the union of IDs.
    all_ids = set(ecg_mapping.keys()) | set(mri_mapping.keys())

    with open(metadata_path, mode="w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["Patient_ID", "MRI_path", "MRI_date", "ECG_path", "ECG_date"])

        for p_id in sorted(all_ids):

            # MRI
            mri_folders = []
            mri_dates = []
            if p_id in mri_mapping:
                for folder, date_obj in mri_mapping[p_id]:
                    mri_folders.append(folder)
                    mri_dates.append(date_obj.strftime("%Y-%m-%d"))

            # ECG
            ecg_folders = []
            ecg_dates = []
            if p_id in ecg_mapping:
                for folder, date_obj in ecg_mapping[p_id]:
                    ecg_folders.append(folder)
                    ecg_dates.append(date_obj.strftime("%Y-%m-%d"))

            writer.writerow([
                p_id,
                ",".join(mri_folders) if mri_folders else "",
                ",".join(mri_dates) if mri_dates else "",
                ",".join(ecg_folders) if ecg_folders else "",
                ",".join(ecg_dates) if ecg_dates else "",
                ECG_states.get(p_id, "")
            ])
        
    # TODO:Pair MRI and ECG (one-to-one), ensure every MRI comes with one ECG, the datatime bwteen them is the closest, and <30 days
    # save to a new csv
    # 4. Pair each MRI with its nearest ECG when the interval is less than 30 days.
    pairs = []
    pairing_results = [] # Used for CSV export.

    for p_id, mris in mri_mapping.items():
        if p_id not in ecg_mapping:
            continue
        
        available_ecgs = ecg_mapping[p_id]
        
        for mri_folder, mri_date in mris:
            best_ecg = None
            min_diff = 31 # Initialize to 31 days.

            for ecg_folder, ecg_date in available_ecgs:
                diff = abs((mri_date - ecg_date).days)
                if min_diff is None:
                    min_diff = diff
                    best_ecg = ecg_folder
                elif diff < min_diff:
                    min_diff = diff
                    best_ecg = ecg_folder
            
            if best_ecg:
                pairs.append((mri_folder, best_ecg))
                pairing_results.append({
                    "Patient_ID": p_id,
                    "Mesh_Folder": mri_folder,
                    "ECG_Folder": best_ecg,
                    "Days_Diff": min_diff
                })

    # 5. Save the pairing results to a new CSV file.
    pd.DataFrame(pairing_results).to_csv(paired_csv_path, index=False)
    print(f"Pairing complete. {len(pairs)} pairs found. Saved to: {paired_csv_path}")
if __name__ == "__main__":
    # digitized_dir = "/path/to/dataset/data_ECG_digitalized"
    # raw_img_dir = "/path/to/dataset/data_ECG"
    # check_ecg_digitization_adaptive(digitized_dir, raw_img_dir)
    # plot_ecg_image(target_folder='patient0008_20170903_1', digitized_dir=digitized_dir, raw_img_dir=raw_img_dir)
    
    # obtain_median_heartbeat_batch()
    
    # VTK_merged_ED_scar()
    
    # cut_ecg_batch()
    # plot_npy_signal()
    # plot_npy_signal_batch()
    
    # pdf_path = r'/path/to/dataset/ECG_series/sample/deidentified.pdf'
    # get_header_datetime_obj(pdf_path)
    
    # CONFIG = {
    # "ecg": "/path/to/dataset/ECG_series",
    # "mri": "/path/to/dataset/DHlab_valid",
    # }
    # metadata_path = "/path/to/dataset/metadata.csv"
    # paired_csv_path = "/path/to/dataset/paired_ecg_mri.csv"
    # paired_ecg_mri_NUH(CONFIG, metadata_path=metadata_path, paired_csv_path=paired_csv_path)
    
    # obtain_median_heartbeat_batch_NUH()
    cut_ecg_batch_NUH()
